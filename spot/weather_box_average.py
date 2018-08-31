#!/usr/bin/python
#
# Module to generate averages for the weather data
# This is not a moving average. Rather, a day is 
# split into bins of AVERAGE_SPAN and the averages are
# computed in each bin for any data that happens to be
# there.

from array import *
import time
import logging
from spot.py_utils import sql_time_str
from openpyxl import Workbook, load_workbook
import os
import datetime
from spot.weatherStatus import path

def test():
    logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',level=logging.DEBUG)
    t0  = round(time.time(),0)+10
    wrec={'timestamp':t0,'barometer':30.12,'insideTemperature':70.1,'insideHumidity':36,'insideDewTemp':36.1,'outsideTemperature':44.7,'outsideHumidity':87,'outsideDewTemp':42.0,'windSpeed':13,'windSpeedAvg':3,'windDirection':132}
    w=weather_box_average(30)
    for i in range(0,40):
        wrec['timestamp'] = t0+2*i
        w.accumulate(wrec)

class weather_box_average(object):
    def __init__(self,AVERAGE_SPAN=30,MailBox=None):
        # averaging span is in seconds
        self.T       = AVERAGE_SPAN
        self.MailBox = MailBox
        self.log     = logging.getLogger('WeatherAveraging')
        self.sum     = array('d',[0,0,0,0,0,0,0,0,0,0,0])
        self.cnt     = array('I',[0,0,0,0,0,0,0,0,0,0,0])
        self.time_ref(time.time())
        self.reset()

    def time_ref(self,t):
        utsec       = t%86400
        self.refday = round(t-utsec,0)
        self.idx    = round(utsec/self.T,0)

    def reset(self):
        for i in range(0,len(self.sum)):
            self.sum[i]=0
            self.cnt[i]=0

    def send(self):
        fname= path + 'DavisVP2_'+datetime.datetime.now().strftime('%m-%d-%y')+'.xlsx'
        print(fname)
        if self.MailBox!=None:
            data = self.avg_wdata
            self.MailBox.put(data)
        if os.path.isfile(fname):
            wb = load_workbook(fname)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(self.avg_wdata.keys()))
        ws.append(list(self.avg_wdata.values()))
        wb.save(fname)


    def accumulate(self,wdata):
        t    = wdata['timestamp']
        vals = array('d',[ wdata['barometer'],
                           wdata['insideTemperature'],
                           wdata['insideHumidity'],
                           wdata['insideDewTemp'],
                           wdata['outsideTemperature'],
                           wdata['outsideHumidity'],
                           wdata['outsideDewTemp'],
                           wdata['windSpeed'],
                           wdata['windSpeedAvg'],
                           wdata['windDirection'],
                           wdata['rainRate']
                           ])
        #self.log.debug("accum " + davis_debug(wdata))
        idx = round((t-self.refday)/self.T,0)
        if (idx!=self.idx):
            # dump a new record
            self.make_record()
            self.time_ref(t)

        # Accumulate the new data
        for i in range(0,len(vals)):
            if ( i!=9 or (i==9 and vals[i]!=0 and vals[i]!=360) ):
                self.sum[i] += vals[i]
                self.cnt[i] += 1

        #self.log.debug("sum   %s P %8.3f T %10.1f R %10.0f Td %10.1f W %10.2f" % ( wdata['timestr'],self.sum[0],self.sum[4],self.sum[5],self.sum[6],self.sum[7]))

    def make_record(self):
        tot_count = 0
        for i in range(0,11):
            if self.cnt[i]>0:
                self.sum[i] /= self.cnt[i]
                tot_count += self.cnt[i]

        t = self.refday + self.idx*self.T

        self.avg_wdata = { 'timestamp'          : t,
                           'timestr'            : sql_time_str(t),
                           'barometer'          : self.sum[0],
                           'insideTemperature'  : self.sum[1],
                           'insideHumidity'     : self.sum[2],
                           'insideDewTemp'      : self.sum[3],
                           'outsideTemperature' : self.sum[4],
                           'outsideHumidity'    : self.sum[5],
                           'outsideDewTemp'     : self.sum[6],
                           'windSpeed'          : self.sum[7],
                           'windSpeedAvg'       : self.sum[8],
                           'windDirection'      : self.sum[9],
                           'rainRate'           : self.sum[10]
                           }

        self.reset()
        if tot_count>0:
            self.send()

    def davis_debug(d):
        return "%s P %6.3f T %4.1f R %3.0f Td %4.1f W %5.2f" % (d['timestr'],d['barometer'],d['outsideTemperature'],d['outsideHumidity'],d['outsideDewTemp'],d['windSpeed'])

if __name__ == "__main__":
    test()
