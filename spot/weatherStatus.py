'''
ATC SpOT Weather Status
Listens to Davis VP2 weather data, logs Dome Health & Safety decisions
'''
from openpyxl import load_workbook
import os
from spot import Dome
import logging
import datetime
from io import BufferedWriter, FileIO
import time
from spot.Dome import Dome
from astral import Astral
path = 'SpOT/data/spot/weather/'

class WeatherStatus(object):
    def __init__(self, timeInterval, keepRunning):
        self.log = logging.getLogger('status_archiver')
        self.logfile = path +'status_' + datetime.datetime.now().strftime('%m-%d-%y') + '.log'
        city_name = 'San Francisco'
        a = Astral()
        a.solar_depression = 'astronomical'
        city = a[city_name]

        timezone = city.timezone
        sun = city.sun(date=datetime.date(2009, 4, 22), local=True)
        self.start = sun['dusk']
        self.stop = sun['dawn']

    #shutdown conditions
    def check_conditions(self, fname):
        dome = Dome()
        status = 0
        pt = datetime.timezone(datetime.timedelta(-1, 61200)) # set timezone PT = -7:00:00 UTC
        now = datetime.datetime.now(tz=pt)

        if self.start < now < self.stop:
            watch_only = False
        else:
            self.log.info(bytes("Watch Only, Dome disconnected\n", encoding='utf-8'))
            watch_only = True
        self.fd.write(bytes(datetime.datetime.now().strftime('%m-%d-%y %H:%M:%S '), encoding='utf-8'))
        if os.path.isfile(fname):
            wb = load_workbook(fname)
            ws = wb.active
            self.data = {
                'Time': datetime.datetime.strptime(ws['B' + str(ws.max_row)].value, '%Y-%m-%d %H:%M:%S'),
                'Barometer': ws['C' + str(ws.max_row)].value,
                'Temp': ws['G' + str(ws.max_row)].value,
                'Humidity': ws['H' + str(ws.max_row)].value,
                'DewTemp': ws['I' + str(ws.max_row)].value,
                'WindSpeed': ws['K' + str(ws.max_row)].value,

            }
            # if self.data['Temp'] > 85: #Test Condition
            #     self.fd.write(bytes("Temperature Red limit (>85): "+str(self.data['Temp'])+'\n', encoding='utf-8'))
            #     self.log.warning("Temperature Red limit")
            #     status += 1
            if self.data['Humidity'] > 85:
                self.fd.write(bytes("Humidity Red limit (> 85%): "+str(self.data['Humidity'])+'\n', encoding='utf-8'))
                self.log.warning("Humidity Red limit")
                status += 1
            if (self.data['Temp']-self.data['DewTemp']) < 2:
                self.fd.write(bytes("Dew Point Red limit (< 2degF): " + str(abs(self.data['DewTemp']-self.data['Temp']))+'\n', encoding='utf-8'))
                self.log.warning("Dew Point Red limit")
                status += 1
            if self.data['WindSpeed'] > 10:
                self.fd.write(bytes("Wind Speed Red limit (> 10mph): "+str(self.data['WindSpeed'])+'\n', encoding='utf-8'))
                self.log.warning("Wind Speed Red limit")
                status += 1
            if status == 0:
                self.fd.write(bytes("Status: Green\n", encoding='utf-8'))
                self.log.info("Status Green")
                if dome.get_dome_state() == b'0' and not watch_only:
                    dome.open()
                    self.fd.write(bytes("Opening Dome\n", encoding='utf-8'))
                #TODO: kickoff image sequencer here
            elif status > 0:
                self.fd.write(bytes("Status: Red\n", encoding='utf-8'))
                self.log.warning("Status Red")
                if dome.get_dome_state() != b'0' and not watch_only:
                    dome.close()
                    self.fd.write(bytes("Closing Dome\n", encoding='utf-8'))

        else:
            self.log.error("No weather Data found %s" %fname)
        dome.disconnect()

    def worker(self, timeInterval, keepWorking):
        #  keeps looking at the queue and empties it to the database
        # keepWorking is an event set in the main thread to allow stoping
        while keepWorking.is_set():
            self.datafile = path + 'DavisVP2_' + datetime.datetime.now().strftime('%m-%d-%y') + '.xlsx'
            self.logfile = path + 'status_' + datetime.datetime.now().strftime('%m-%d-%y') + '.log'
            try:
                if os.path.exists(self.logfile):
                    self.fd = BufferedWriter(FileIO(self.logfile, 'ab'))
                else:
                    self.fd = BufferedWriter(FileIO(self.logfile, 'wb'))

            except IOError:
                self.log.error("Cannot create log file %s" % self.logfile)
            try:
                if os.path.exists(self.datafile):
                    self.fd = BufferedWriter(FileIO(self.logfile, 'ab'))
                else:
                    self.fd = BufferedWriter(FileIO(self.logfile, 'wb'))

            except IOError:
                self.log.error("Cannot create data file %s" % self.datafile)

            self.check_conditions(self.datafile)
            time.sleep(timeInterval)

